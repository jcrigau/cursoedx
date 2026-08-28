"""Exportación de las novedades en el formato que espera el liquidador.

Las columnas son las del formulario que la escuela usa hoy, en el mismo orden,
para que la planilla se pueda copiar y pegar sin reacomodar nada. Si otra
institución usa otras columnas, se cambia acá: el resto del sistema no se
entera.
"""

import csv
import io

from .models import Destino, Novedad

# Encabezados tal como figuran en la planilla del liquidador.
COLUMNAS = [
    "Marca temporal",
    "Nivel - Área",
    "Novedad",
    "Fecha",
    "Espacio Curricular",
    "Apellido y Nombre",
    "Motivo",
    "Presenta Certificado?",
    "Reemplazante",
    "Jornada Completa",
    "Cantidad de Horas",
    "Tiempo Determinado",
    "Fecha de Finalización",
    "Planilla",
    "OBSERVACIONES",
]


def _si_no(valor: bool) -> str:
    return "Sí" if valor else "No"


def _fecha(valor) -> str:
    return valor.strftime("%d/%m/%Y") if valor else ""


def fila_de(novedad: Novedad) -> list:
    """Una novedad convertida en la fila que espera la planilla."""
    return [
        novedad.creado_en.strftime("%d/%m/%Y %H:%M"),
        novedad.nivel,
        novedad.get_tipo_display(),
        _fecha(novedad.fecha),
        novedad.espacio,
        novedad.legajo.nombre_completo,
        novedad.motivo,
        _si_no(novedad.presenta_certificado),
        novedad.reemplazante,
        _si_no(novedad.jornada_completa),
        novedad.horas if novedad.horas else (novedad.dias or ""),
        _si_no(novedad.tiempo_determinado),
        _fecha(novedad.fecha_fin),
        novedad.get_destino_display().replace("Planilla ", ""),
        novedad.observaciones,
    ]


def novedades_a_exportar(periodo, destino: str | None = None, solo_a_informar: bool = True):
    """Novedades del período, filtradas como se las va a informar."""
    consulta = periodo.novedades.select_related("legajo", "cargo__nivel")
    if solo_a_informar:
        consulta = consulta.filter(impacta_haberes=True)
    if destino:
        consulta = consulta.filter(destino=destino)
    return consulta.order_by("destino", "legajo__apellido", "fecha")


def a_csv(periodo, destino: str | None = None, solo_a_informar: bool = True) -> str:
    """CSV listo para pegar en la planilla compartida."""
    salida = io.StringIO()
    escritor = csv.writer(salida)
    escritor.writerow(COLUMNAS)
    for novedad in novedades_a_exportar(periodo, destino, solo_a_informar):
        escritor.writerow(fila_de(novedad))
    return salida.getvalue()


def a_xlsx(periodo, solo_a_informar: bool = True) -> bytes:
    """Excel con una hoja por planilla (Oficial e Interna)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = Workbook()
    libro.remove(libro.active)

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="16232E")

    for destino, etiqueta in Destino.choices:
        hoja = libro.create_sheet(etiqueta.replace("Planilla ", ""))
        hoja.append(COLUMNAS)
        for celda in hoja[1]:
            celda.font = encabezado
            celda.fill = fondo
            celda.alignment = Alignment(vertical="center", wrap_text=True)

        for novedad in novedades_a_exportar(periodo, destino, solo_a_informar):
            hoja.append(fila_de(novedad))

        anchos = [18, 12, 14, 12, 24, 26, 26, 10, 24, 10, 10, 10, 14, 10, 40]
        for indice, ancho in enumerate(anchos, start=1):
            hoja.column_dimensions[get_column_letter(indice)].width = ancho
        hoja.freeze_panes = "A2"

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def resumen_por_persona(periodo, solo_a_informar: bool = True) -> list[dict]:
    """Agrupa las novedades por persona: es como se revisa antes de cerrar."""
    agrupadas: dict[int, dict] = {}
    for novedad in novedades_a_exportar(periodo, solo_a_informar=solo_a_informar):
        fila = agrupadas.setdefault(
            novedad.legajo_id,
            {"legajo": novedad.legajo, "novedades": [], "destinos": set()},
        )
        fila["novedades"].append(novedad)
        fila["destinos"].add(novedad.get_destino_display().replace("Planilla ", ""))
    return sorted(agrupadas.values(), key=lambda fila: fila["legajo"].nombre_completo)
