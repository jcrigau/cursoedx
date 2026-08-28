"""Exportación en el formato del liquidador, y las pantallas del mes."""

import csv
import io
from datetime import date

import pytest

from core.models import Membresia, Rol, Usuario
from legajos.models import FuentePago
from novedades.compilador import compilar
from novedades.exportar import COLUMNAS, a_csv, a_xlsx, resumen_por_persona
from novedades.models import Destino, EstadoPeriodo, PeriodoNovedades

from .conftest import ANIO, MES, dar_cargo, dar_licencia, dar_materia


@pytest.fixture
def periodo_compilado(institucion, periodo, docente, nivel):
    matematica = dar_materia(institucion, nivel, "Matemática")
    dar_cargo(
        institucion, docente, alta=date(ANIO, MES, 4), materia=matematica, horas=5, nivel=nivel
    )
    dar_licencia(
        institucion,
        docente,
        date(ANIO, MES, 10),
        date(ANIO, MES, 14),
        nombre="Sin goce",
        con_goce=False,
    )
    compilar(periodo)
    return periodo


class TestExportacion:
    def test_el_csv_tiene_las_columnas_de_la_planilla(self, periodo_compilado):
        lector = csv.reader(io.StringIO(a_csv(periodo_compilado)))
        assert next(lector) == COLUMNAS

    def test_cada_novedad_es_una_fila(self, periodo_compilado):
        filas = list(csv.reader(io.StringIO(a_csv(periodo_compilado))))
        assert len(filas) == 1 + periodo_compilado.novedades.filter(impacta_haberes=True).count()

    def test_los_datos_salen_en_el_formato_esperado(self, periodo_compilado, docente):
        filas = list(csv.DictReader(io.StringIO(a_csv(periodo_compilado))))
        alta = next(fila for fila in filas if fila["Novedad"] == "Alta")

        assert alta["Apellido y Nombre"] == docente.nombre_completo
        assert alta["Fecha"] == "04/05/2026"
        assert alta["Espacio Curricular"] == "Matemática"
        assert alta["Planilla"] == "Oficial"
        assert alta["Presenta Certificado?"] in {"Sí", "No"}

    def test_se_puede_exportar_una_sola_planilla(self, institucion, periodo, docente, nivel):
        dar_cargo(institucion, docente, alta=date(ANIO, MES, 4), nivel=nivel)
        otro_legajo = docente.__class__.objects.create(
            institucion=institucion, apellido="Paz", nombre="Ana", cuil="27-31456789-3"
        )
        dar_cargo(
            institucion,
            otro_legajo,
            alta=date(ANIO, MES, 5),
            fuente=FuentePago.INTERNO,
            nivel=nivel,
        )
        compilar(periodo)

        filas = list(csv.DictReader(io.StringIO(a_csv(periodo, destino=Destino.OFICIAL))))
        assert len(filas) == 1
        assert filas[0]["Planilla"] == "Oficial"

    def test_no_exporta_lo_que_no_impacta(self, institucion, periodo, docente, nivel):
        dar_cargo(institucion, docente, nivel=nivel)
        dar_licencia(institucion, docente, date(ANIO, MES, 4), date(ANIO, MES, 6), con_goce=True)
        compilar(periodo)

        filas = list(csv.reader(io.StringIO(a_csv(periodo))))
        assert len(filas) == 1  # solo el encabezado

    def test_el_excel_trae_una_hoja_por_planilla(self, periodo_compilado):
        from openpyxl import load_workbook

        libro = load_workbook(io.BytesIO(a_xlsx(periodo_compilado)))
        assert libro.sheetnames == ["Oficial", "Interna"]
        assert [celda.value for celda in libro["Oficial"][1]] == COLUMNAS

    def test_agrupa_por_persona_para_revisar(self, periodo_compilado, docente):
        agrupadas = resumen_por_persona(periodo_compilado)
        assert len(agrupadas) == 1
        assert agrupadas[0]["legajo"] == docente
        assert len(agrupadas[0]["novedades"]) == 2


@pytest.mark.django_db
class TestPantallas:
    def test_exige_login(self, client):
        respuesta = client.get("/novedades/")
        assert respuesta.status_code == 302

    def test_lista_los_periodos(self, client, secretaria, periodo_compilado):
        client.force_login(secretaria)
        respuesta = client.get("/novedades/")
        assert respuesta.status_code == 200
        assert len(respuesta.context["periodos"]) == 1

    def test_el_detalle_agrupa_por_persona(self, client, secretaria, periodo_compilado):
        client.force_login(secretaria)
        respuesta = client.get(f"/novedades/{ANIO}/{MES}/")
        assert respuesta.status_code == 200
        assert respuesta.context["resumen"]["a_informar"] == 2

    def test_compilar_desde_la_pantalla(
        self, client, secretaria, institucion, periodo, docente, nivel
    ):
        dar_cargo(institucion, docente, alta=date(ANIO, MES, 4), nivel=nivel)
        client.force_login(secretaria)
        client.post(f"/novedades/{ANIO}/{MES}/", {"accion": "compilar"})
        assert periodo.novedades.count() == 1

    def test_marcar_informadas(self, client, secretaria, periodo_compilado):
        client.force_login(secretaria)
        ids = list(periodo_compilado.novedades.values_list("pk", flat=True))
        client.post(
            f"/novedades/{ANIO}/{MES}/", {"accion": "informadas", "novedad": [str(i) for i in ids]}
        )
        assert periodo_compilado.novedades.filter(informada=True).count() == len(ids)

    def test_cerrar_pide_confirmacion_si_falta_informar(
        self, client, secretaria, periodo_compilado
    ):
        client.force_login(secretaria)
        client.post(f"/novedades/{ANIO}/{MES}/", {"accion": "cerrar"})

        periodo_compilado.refresh_from_db()
        assert not periodo_compilado.esta_cerrado  # pidió confirmar primero

        client.post(f"/novedades/{ANIO}/{MES}/", {"accion": "cerrar", "confirmar": "1"})
        periodo_compilado.refresh_from_db()
        assert periodo_compilado.esta_cerrado

    def test_reabrir_exige_motivo(self, client, secretaria, periodo_compilado):
        periodo_compilado.cerrar(usuario=secretaria)
        client.force_login(secretaria)

        client.post(f"/novedades/{ANIO}/{MES}/", {"accion": "reabrir", "motivo": ""})
        periodo_compilado.refresh_from_db()
        assert periodo_compilado.esta_cerrado

        client.post(
            f"/novedades/{ANIO}/{MES}/", {"accion": "reabrir", "motivo": "Faltó una licencia"}
        )
        periodo_compilado.refresh_from_db()
        assert periodo_compilado.estado == EstadoPeriodo.REABIERTO

    def test_descarga_el_excel(self, client, secretaria, periodo_compilado):
        client.force_login(secretaria)
        respuesta = client.get(f"/novedades/{ANIO}/{MES}/exportar/?formato=xlsx")
        assert respuesta.status_code == 200
        assert "spreadsheetml" in respuesta["Content-Type"]
        assert "novedades-2026-05.xlsx" in respuesta["Content-Disposition"]

    def test_descarga_el_pdf(self, client, secretaria, periodo_compilado):
        client.force_login(secretaria)
        respuesta = client.get(f"/novedades/{ANIO}/{MES}/exportar/?formato=pdf")
        assert respuesta.content.startswith(b"%PDF")

    def test_la_exportacion_queda_auditada(self, client, secretaria, periodo_compilado):
        from core.models import RegistroAuditoria

        client.force_login(secretaria)
        client.get(f"/novedades/{ANIO}/{MES}/exportar/?formato=csv")

        registro = RegistroAuditoria.objects.filter(accion="EXPORTACION").first()
        assert registro is not None
        assert "Novedades" in registro.descripcion or "novedades" in registro.descripcion


@pytest.mark.django_db
class TestAccesoDelLiquidador:
    @pytest.fixture
    def liquidador(self, institucion):
        usuario = Usuario.objects.create_user(
            email="contador@estudio.com", password="x", nombre="Carlos", apellido="Ruiz"
        )
        Membresia.objects.create(usuario=usuario, institucion=institucion, rol=Rol.LIQUIDADOR)
        return usuario

    def test_no_ve_un_periodo_abierto(self, client, liquidador, periodo_compilado):
        """Un borrador puede cambiar: no se le muestra hasta cerrarlo."""
        client.force_login(liquidador)
        assert client.get(f"/novedades/{ANIO}/{MES}/").status_code == 403

    def test_ve_el_periodo_cerrado(self, client, liquidador, periodo_compilado, secretaria):
        periodo_compilado.cerrar(usuario=secretaria)
        client.force_login(liquidador)
        respuesta = client.get(f"/novedades/{ANIO}/{MES}/")
        assert respuesta.status_code == 200
        assert respuesta.context["puede_editar"] is False

    def test_descarga_el_paquete_cerrado(self, client, liquidador, periodo_compilado, secretaria):
        periodo_compilado.cerrar(usuario=secretaria)
        client.force_login(liquidador)
        respuesta = client.get(f"/novedades/{ANIO}/{MES}/exportar/?formato=xlsx")
        assert respuesta.status_code == 200

    def test_no_puede_compilar_ni_cerrar(self, client, liquidador, periodo_compilado):
        client.force_login(liquidador)
        respuesta = client.post(f"/novedades/{ANIO}/{MES}/", {"accion": "compilar"})
        assert respuesta.status_code == 403

    def test_no_ve_las_novedades_de_otra_escuela(
        self, client, liquidador, otra_institucion, secretaria
    ):
        ajeno = PeriodoNovedades.objects.create(
            institucion=otra_institucion, anio=ANIO, mes=MES, estado=EstadoPeriodo.CERRADO
        )
        client.force_login(liquidador)
        respuesta = client.get(f"/novedades/{ajeno.anio}/{ajeno.mes}/")
        # Ve el de su propia institución (vacío), nunca el de la otra.
        assert respuesta.status_code in {403, 200}
        if respuesta.status_code == 200:
            assert respuesta.context["periodo"].institucion != otra_institucion
