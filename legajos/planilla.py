"""El personal en Excel: se baja, se corrige, se vuelve a subir.

Una planta de 130 personas no se tipea en formularios. El circuito es el que
la secretaría ya conoce de las planillas: se descarga la lista actual, se
corrige o se agregan filas en Excel, y se sube. Lo que ya existe se
actualiza, lo nuevo se crea, y lo dudoso se observa sin tocar nada.

La identidad es el CUIL: es el único dato que no cambia. Pero una escuela que
recién arranca tiene la lista de apellidos mucho antes que los CUIL, así que
cuando la columna viene vacía se identifica por apellido y nombre y el CUIL se
completa después, en cuanto aparezca. Un CUIL *mal escrito* es otra cosa: eso
se saltea, porque termina en la liquidación.

Los cargos van en su propia hoja de la planilla de carga (``importar_cargos``)
y no en esta: un cargo tiene materia, curso, fuente de pago y situación de
revista, y eso mal importado son errores de liquidación. Por eso cada fila pasa
por ``full_clean`` antes de guardarse y lo que no valida queda observado en vez
de entrar a medias.
"""

import re
from datetime import date, datetime

from django.core.exceptions import ValidationError

from core.planillas import MARCA_EJEMPLO, Resultado, es_ejemplo
from estructura.models import Materia

from .models import EstadoLegajo, Legajo, Plantel

ENCABEZADOS = [
    "CUIL",
    "Apellido",
    "Nombre",
    "DNI",
    "Email",
    "Teléfono",
    "Fecha de ingreso",
    "Fecha de nacimiento",
    "Obra social",
    "Domicilio",
    "Localidad",
    "Estado",
    "Plantel",
    "Materias que puede dar",
]

SEPARADOR_MATERIAS = " | "

# La planilla vacía sale con una fila de muestra: es lo que evita que alguien
# tipee el CUIL sin guiones o invente el nombre del plantel. Va marcada, así
# que si vuelve sin borrar no crea a nadie (ver ``core.planillas``).
EJEMPLOS = [
    [
        f"{MARCA_EJEMPLO} 27-30111222-4",
        "Benítez",
        "María Laura",
        "30111222",
        "mlbenitez@gmail.com",
        "2657 55-1234",
        "01/03/2018",
        "14/06/1985",
        "OSDE",
        "San Martín 456",
        "Villa Mercedes",
        "Activo",
        "Docente",
        "Matemática | Física",
    ],
    [
        f"{MARCA_EJEMPLO} 20-25888777-1",
        "Ferreyra",
        "Hugo Ramón",
        "25888777",
        "",
        "2657 55-9876",
        "10/04/2021",
        "",
        "",
        "",
        "Villa Mercedes",
        "Activo",
        "Maestranza / ordenanza",
        "",
    ],
]


def exportar(institucion):
    """La planta completa, lista para corregir y volver a subir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Personal"

    hoja.append(ENCABEZADOS)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F6B63")

    personas = (
        Legajo.objects.filter(institucion=institucion)
        .prefetch_related("materias_que_puede_dar")
        .order_by("apellido", "nombre")
    )
    for legajo in personas:
        hoja.append(
            [
                legajo.cuil,
                legajo.apellido,
                legajo.nombre,
                legajo.dni,
                legajo.email,
                legajo.telefono,
                legajo.fecha_ingreso,
                legajo.fecha_nacimiento,
                legajo.obra_social,
                legajo.domicilio,
                legajo.localidad,
                legajo.get_estado_display(),
                legajo.get_plantel_display(),
                SEPARADOR_MATERIAS.join(
                    materia.nombre for materia in legajo.materias_que_puede_dar.all()
                ),
            ]
        )

    if not personas:
        for ejemplo in EJEMPLOS:
            hoja.append(ejemplo)
            for celda in hoja[hoja.max_row]:
                celda.font = Font(italic=True, color="8A6D3B")
                celda.fill = PatternFill("solid", fgColor="FFF8E1")

    anchos = [22, 22, 22, 12, 32, 16, 14, 14, 18, 26, 16, 10, 18, 40]
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = "A2"
    return libro


def importar(institucion, archivo) -> Resultado:
    """Lee la planilla y aplica lo que se pueda; lo dudoso se observa.

    No borra a nadie: una fila que falta en el archivo no es una baja, es una
    fila que falta. Las bajas se hacen en el sistema, con su motivo.
    """
    from openpyxl import load_workbook

    resultado = Resultado("Personal")
    try:
        libro = load_workbook(archivo, data_only=True)
    except Exception:
        resultado.observaciones.append(
            "No se pudo leer el archivo. Tiene que ser el Excel descargado de acá "
            "(.xlsx), no un CSV ni una planilla de otro formato."
        )
        return resultado

    # Sirve tanto el Excel que se baja de acá —una hoja, encabezado arriba de
    # todo— como la hoja «Personal» de la planilla de carga, que trae dos filas
    # de ayuda antes. Buscar el encabezado es más barato que pedir dos formatos.
    hoja = libro["Personal"] if "Personal" in libro.sheetnames else libro.active
    encabezado = _fila_del_encabezado(hoja)

    materias = {
        _clave(materia.nombre): materia
        for materia in Materia.objects.filter(institucion=institucion)
    }

    indice = _indice_por_nombre(institucion)
    filas = hoja.iter_rows(min_row=encabezado + 1, values_only=True)
    for numero, fila in enumerate(filas, start=encabezado + 1):
        if fila is None or all(celda in (None, "") for celda in fila):
            continue
        if es_ejemplo(fila[0]):
            continue
        fila = list(fila) + [None] * (len(ENCABEZADOS) - len(fila))

        apellido = str(fila[1] or "").strip()
        nombre = str(fila[2] or "").strip()
        if not apellido or not nombre:
            resultado.observaciones.append(
                f"Fila {numero}: falta el apellido o el nombre. Se salteó."
            )
            continue

        cuil = _normalizar_cuil(fila[0])
        if cuil is None and str(fila[0] or "").strip():
            resultado.observaciones.append(
                f"Fila {numero} ({apellido}): el CUIL «{fila[0]}» no tiene 11 dígitos. "
                "Se salteó la fila: un CUIL mal cargado termina en la liquidación."
            )
            continue

        legajo, duda = _quien_es(institucion, cuil, apellido, nombre, indice)
        if duda:
            resultado.observaciones.append(f"Fila {numero} ({apellido}, {nombre}): {duda}")
            continue

        ingreso = _como_fecha(fila[6])
        nuevo = legajo is None
        if nuevo:
            legajo = Legajo(institucion=institucion, cuil=cuil or "", fecha_ingreso=ingreso)
            if ingreso is None:
                resultado.avisar(
                    "personas cargadas sin fecha de ingreso: hasta completarla no se les "
                    "puede computar la antigüedad"
                )
        elif cuil and not legajo.cuil:
            legajo.cuil = cuil  # llegó el dato que faltaba

        legajo.apellido = apellido
        legajo.nombre = nombre
        legajo.dni = str(fila[3] or "").strip()
        legajo.email = str(fila[4] or "").strip()
        legajo.telefono = str(fila[5] or "").strip()
        if ingreso is not None:
            legajo.fecha_ingreso = ingreso
        legajo.fecha_nacimiento = _como_fecha(fila[7])
        legajo.obra_social = str(fila[8] or "").strip()
        legajo.domicilio = str(fila[9] or "").strip()
        legajo.localidad = str(fila[10] or "").strip()

        estado = _clave(str(fila[11] or ""))
        if estado.startswith("baja") or estado == "de baja":
            legajo.estado = EstadoLegajo.BAJA
        elif estado:
            legajo.estado = EstadoLegajo.ACTIVO

        plantel = _plantel_de(fila[12])
        if plantel is not None:
            legajo.plantel = plantel
        elif fila[12] not in (None, ""):
            resultado.observaciones.append(
                f"Fila {numero} ({apellido}): plantel «{fila[12]}» no reconocido; quedó "
                f"«{legajo.get_plantel_display()}». Vale: docente, preceptor, directivo, "
                "administrativo o maestranza."
            )

        legajo.save()

        elegidas, desconocidas = _materias_de(fila[13], materias)
        legajo.materias_que_puede_dar.set(elegidas)
        for nombre_materia in desconocidas:
            resultado.observaciones.append(
                f"Fila {numero} ({apellido}): la materia «{nombre_materia}» no existe "
                "en la escuela; se ignoró. Cargala primero en Materias."
            )

        if nuevo:
            resultado.creados += 1
            indice.setdefault(_clave_persona(apellido, nombre), []).append(legajo)
        else:
            resultado.actualizados += 1

    return resultado


def _fila_del_encabezado(hoja, hasta: int = 6) -> int:
    """En qué fila están los títulos de las columnas."""
    for numero in range(1, hasta + 1):
        if _clave(str(hoja.cell(row=numero, column=1).value or "")).lower() == "cuil":
            return numero
    return 1


def _quien_es(institucion, cuil, apellido, nombre, indice):
    """De quién es esta fila: el CUIL manda; si no está, el apellido y el nombre.

    Identificar por nombre es peor que por CUIL —hay homónimos—, así que cuando
    el nombre alcanza para dos personas no se toca ninguna: se avisa y que lo
    resuelva alguien que sepa cuál es cuál.
    """
    if cuil:
        propio = Legajo.objects.filter(institucion=institucion, cuil=cuil).first()
        if propio is not None:
            return propio, None

    candidatos = indice.get(_clave_persona(apellido, nombre), [])
    if cuil:
        # El que ya tiene otro CUIL es otra persona con el mismo nombre.
        candidatos = [legajo for legajo in candidatos if not legajo.cuil]
    if len(candidatos) > 1:
        return None, (
            "hay más de una persona con ese apellido y nombre, y la fila no trae CUIL "
            "para distinguirlas. No se tocó ninguna."
        )
    if candidatos:
        return candidatos[0], None
    return None, None


def _indice_por_nombre(institucion) -> dict[str, list]:
    indice: dict[str, list] = {}
    for legajo in Legajo.objects.filter(institucion=institucion):
        indice.setdefault(_clave_persona(legajo.apellido, legajo.nombre), []).append(legajo)
    return indice


def _clave_persona(apellido, nombre) -> str:
    """Apellido y nombre comparables: sin tildes, sin mayúsculas, sin dobles espacios."""
    return " ".join(_clave(f"{apellido} {nombre}").lower().split())


def _clave(texto: str) -> str:
    from core.texto import sin_tildes

    return sin_tildes(texto).strip()


def _normalizar_cuil(crudo) -> str | None:
    """Acepta el CUIL como venga —con guiones, como número de Excel— o nada."""
    if crudo is None:
        return None
    digitos = re.sub(r"\D", "", str(crudo))
    if len(digitos) != 11:
        return None
    return f"{digitos[:2]}-{digitos[2:10]}-{digitos[10]}"


def _plantel_de(crudo) -> str | None:
    """El plantel de la celda, escrito como sea: «Preceptora», «ordenanza»…"""
    clave = _clave(str(crudo or ""))
    if not clave:
        return None
    equivalencias = {
        "docente": Plantel.DOCENTE,
        "profesor": Plantel.DOCENTE,
        "precept": Plantel.PRECEPTOR,
        "directiv": Plantel.DIRECTIVO,
        "administrat": Plantel.ADMINISTRATIVO,
        "secretari": Plantel.ADMINISTRATIVO,
        "maestranza": Plantel.MAESTRANZA,
        "ordenanza": Plantel.MAESTRANZA,
    }
    for prefijo, valor in equivalencias.items():
        if clave.startswith(prefijo):
            return valor
    return None


def _como_fecha(crudo):
    """Excel devuelve datetime; una edición a mano puede dejar texto."""
    if crudo in (None, ""):
        return None
    if isinstance(crudo, datetime):
        return crudo.date()
    if isinstance(crudo, date):
        return crudo
    for formato in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(crudo).strip(), formato).date()
        except ValueError:
            continue
    return None


def _materias_de(crudo, materias: dict) -> tuple[list, list[str]]:
    """Las materias de la celda, separadas por | o por coma."""
    if not crudo:
        return [], []
    encontradas, desconocidas = [], []
    for pedazo in re.split(r"[|,;]", str(crudo)):
        nombre = pedazo.strip()
        if not nombre:
            continue
        materia = materias.get(_clave(nombre))
        if materia is None:
            desconocidas.append(nombre)
        else:
            encontradas.append(materia)
    return encontradas, desconocidas


# ---------------------------------------------------------------------------
# Los cargos
# ---------------------------------------------------------------------------
#
# La clave de un cargo no es un número: es *qué hace la persona y quién se lo
# paga*. Dos filas con la misma persona, la misma materia, el mismo curso y la
# misma fuente son el mismo cargo, aunque cambien las horas. Con eso alcanza
# para que volver a subir la planilla corrija en vez de duplicar.


def importar_cargos(institucion, libro, ciclo=None) -> Resultado:
    """La hoja «Cargos» de la planilla de carga.

    Cada fila pasa por ``full_clean``: un cargo de horas cátedra sin materia, o
    una baja sin motivo, queda observado y no entra. Es plata de por medio.
    """
    from core.planillas import entero, fecha, leer, opcion, opciones_de, si_no, texto
    from estructura.models import CicloLectivo, Curso, Nivel, TipoNivel
    from estructura.planilla import buscar_curso

    from .models import Cargo, FuentePago, MotivoBaja, SituacionRevista, TipoCargo

    resultado = Resultado("Cargos")
    if ciclo is None:
        ciclo = CicloLectivo.objects.filter(institucion=institucion).order_by("-anio").first()
    indice = _indice_por_nombre(institucion)
    alta_por_omision = ciclo.fecha_inicio if ciclo else date.today()
    hay_cursos = bool(ciclo) and Curso.objects.filter(ciclo_lectivo=ciclo).exists()
    sin_cursos = 0
    # Dos filas que caen en el mismo cargo no se fusionan calladas: una que se
    # pierde son horas que nadie cobra. Dentro de una misma corrida, la segunda
    # se observa; entre corridas, la clave sigue sirviendo para actualizar.
    vistos: dict[tuple, int] = {}

    for numero, fila in leer(libro, "Cargos"):
        legajo, duda = _quien_es(
            institucion,
            _normalizar_cuil(fila.get("CUIL")),
            texto(fila.get("Apellido")),
            texto(fila.get("Nombre")),
            indice,
        )
        if duda:
            resultado.observar(numero, duda)
            continue
        if legajo is None:
            resultado.observar(
                numero,
                "no hay nadie con ese CUIL ni con ese apellido y nombre. Cargá primero la "
                "hoja Personal.",
            )
            continue

        tipo = opcion(fila.get("Tipo de cargo"), TipoCargo.choices)
        if tipo is None:
            resultado.observar(
                numero,
                f"«{texto(fila.get('Tipo de cargo'))}» no es un tipo de cargo. "
                f"Vale {opciones_de(TipoCargo.choices)}.",
            )
            continue
        situacion = opcion(fila.get("Situación de revista"), SituacionRevista.choices)
        fuente = opcion(fila.get("Fuente de pago"), FuentePago.choices)
        if situacion is None or fuente is None:
            resultado.observar(
                numero,
                "falta la situación de revista o la fuente de pago. La fuente es la que "
                "decide a qué planilla van las novedades, así que no se puede suponer.",
            )
            continue

        nivel = None
        if texto(fila.get("Nivel")):
            tipo_nivel = opcion(fila.get("Nivel"), TipoNivel.choices)
            nivel = Nivel.objects.filter(institucion=institucion, tipo=tipo_nivel).first()

        materia = None
        if texto(fila.get("Materia")):
            materia = _materia_del_cargo(institucion, nivel, fila.get("Materia"))
            if materia is None:
                resultado.observar(
                    numero,
                    f"la materia «{texto(fila.get('Materia'))}» no existe. Cargala primero "
                    "en la hoja Materias, escrita igual.",
                )
                continue
            nivel = nivel or materia.nivel

        curso = None
        if texto(fila.get("Curso")):
            if not hay_cursos:
                sin_cursos += 1
                continue
            curso, problema = buscar_curso(institucion, ciclo, fila.get("Curso"), nivel)
            if problema:
                resultado.observar(numero, problema)
                continue

        alta = fecha(fila.get("Fecha de alta"))
        if alta is None:
            alta = alta_por_omision
            resultado.avisar(f"cargos sin fecha de alta: se usó {alta_por_omision:%d/%m/%Y}")

        denominacion = texto(fila.get("Denominación"))[:120]
        firma = (
            legajo.pk,
            tipo,
            materia.pk if materia else None,
            curso.pk if curso else None,
            fuente,
            denominacion,
        )
        if firma in vistos:
            resultado.observar(
                numero,
                f"no se distingue de la fila {vistos[firma]}: misma persona, mismo tipo, "
                "misma materia, mismo curso y misma fuente de pago. Si son dos cargos "
                "distintos hay que diferenciarlos —el curso, o la denominación—; si es "
                "el mismo, sobra una fila.",
            )
            continue
        vistos[firma] = numero

        cargo = Cargo.objects.filter(
            institucion=institucion,
            legajo=legajo,
            tipo=tipo,
            materia=materia,
            curso=curso,
            fuente_pago=fuente,
            denominacion=denominacion,
        ).first()
        creado = cargo is None
        if creado:
            cargo = Cargo(
                institucion=institucion,
                legajo=legajo,
                tipo=tipo,
                materia=materia,
                curso=curso,
                fuente_pago=fuente,
                denominacion=denominacion,
            )

        cargo.nivel = nivel
        cargo.horas_semanales = entero(fila.get("Horas semanales"))
        cargo.jornada_completa = si_no(fila.get("Jornada completa"))
        cargo.situacion_revista = situacion
        cargo.fecha_alta = alta
        cargo.fecha_baja = fecha(fila.get("Fecha de baja"))
        cargo.motivo_baja = opcion(fila.get("Motivo de baja"), MotivoBaja.choices, defecto="") or ""
        cargo.resolucion_numero = texto(fila.get("Resolución n°"))[:50]
        cargo.resolucion_fecha = fecha(fila.get("Fecha de resolución"))
        cargo.observaciones = texto(fila.get("Observaciones"))

        try:
            cargo.full_clean(exclude=["resolucion_archivo"])
        except ValidationError as error:
            resultado.observar(numero, _en_una_linea(error, Cargo))
            continue
        cargo.save()
        if creado:
            resultado.creados += 1
        else:
            resultado.actualizados += 1

    if sin_cursos:
        resultado.problema(
            f"{sin_cursos} cargos están asignados a un curso y la escuela todavía no "
            "tiene cursos cargados. Completá «Ciclo y períodos», «Turnos» y «Grilla "
            "horaria» —de ahí salen los cursos— y volvé a correr el comando."
        )
    return resultado


def _materia_del_cargo(institucion, nivel, escrito):
    from core.planillas import clave as comparable
    from estructura.models import Materia

    buscado = comparable(escrito)
    candidatas = Materia.objects.filter(institucion=institucion)
    if nivel is not None:
        candidatas = candidatas.filter(nivel=nivel)
    for materia in candidatas:
        if comparable(materia.nombre) == buscado:
            return materia
    return None


def _en_una_linea(error: ValidationError, modelo=None) -> str:
    """El error de validación como lo leería la secretaría, no el programador.

    Los campos se nombran con su etiqueta («horas semanales»), no con el
    nombre en la base («horas_semanales»): el informe lo lee quien completó
    el Excel.
    """
    partes = []
    for campo, mensajes in getattr(error, "message_dict", {}).items():
        etiqueta = ""
        if campo != "__all__":
            etiqueta = f"{_como_se_llama(modelo, campo)}: "
        partes.append(etiqueta + " ".join(mensajes))
    return " ".join(partes) or str(error)


def _como_se_llama(modelo, campo: str) -> str:
    from django.core.exceptions import FieldDoesNotExist

    if modelo is None:
        return campo.replace("_", " ")
    try:
        return str(modelo._meta.get_field(campo).verbose_name)
    except FieldDoesNotExist:
        return campo.replace("_", " ")
