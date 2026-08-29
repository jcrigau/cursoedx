"""La planilla del personal: bajar, corregir, subir.

Lo que importa: el ida y vuelta no pierde nada, el CUIL identifica, lo dudoso
se observa sin tocar, y volver a subir no duplica.
"""

from datetime import date
from io import BytesIO

import pytest
from django.urls import reverse

from estructura.models import Materia, Nivel, TipoNivel
from legajos import planilla
from legajos.models import EstadoLegajo, Legajo


@pytest.fixture
def escuela_con_gente(institucion):
    nivel = Nivel.objects.create(institucion=institucion, tipo=TipoNivel.SECUNDARIO)
    quimica = Materia.objects.create(institucion=institucion, nivel=nivel, nombre="Química")
    persona = Legajo.objects.create(
        institucion=institucion,
        apellido="Benítez",
        nombre="Ana",
        cuil="27-30000001-1",
        email="ana@x.com",
        fecha_ingreso=date(2020, 3, 1),
    )
    persona.materias_que_puede_dar.add(quimica)
    return {"institucion": institucion, "persona": persona, "quimica": quimica}


def como_archivo(libro) -> BytesIO:
    contenido = BytesIO()
    libro.save(contenido)
    contenido.seek(0)
    return contenido


class TestElIdaYVuelta:
    def test_exportar_y_reimportar_no_cambia_nada(self, escuela_con_gente):
        archivo = como_archivo(planilla.exportar(escuela_con_gente["institucion"]))

        resultado = planilla.importar(escuela_con_gente["institucion"], archivo)

        assert resultado.creados == 0
        assert resultado.actualizados == 1
        assert not resultado.observaciones
        persona = Legajo.objects.get(pk=escuela_con_gente["persona"].pk)
        assert persona.email == "ana@x.com"
        assert list(persona.materias_que_puede_dar.all()) == [escuela_con_gente["quimica"]]

    def test_una_correccion_en_el_excel_se_aplica(self, escuela_con_gente):
        libro = planilla.exportar(escuela_con_gente["institucion"])
        hoja = libro.active
        hoja.cell(row=2, column=5, value="nuevo@correo.com")  # email

        planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert Legajo.objects.get(pk=escuela_con_gente["persona"].pk).email == "nuevo@correo.com"

    def test_una_fila_nueva_crea_a_la_persona(self, escuela_con_gente):
        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.append(
            ["20-11222333-4", "Nueva", "Persona", "", "n@x.com", "", date(2026, 3, 1)]
        )

        resultado = planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert resultado.creados == 1
        assert Legajo.objects.filter(cuil="20-11222333-4", apellido="Nueva").exists()

    def test_el_cuil_se_normaliza_como_venga(self, escuela_con_gente):
        """Excel suele comerse los guiones y devolver el CUIL como número."""
        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.append([20112223334, "Numérica", "Fila", "", "", "", date(2026, 3, 1)])

        planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert Legajo.objects.filter(cuil="20-11222333-4").exists()


class TestLoQueSeObserva:
    def test_cuil_invalido_se_saltea_y_se_explica(self, escuela_con_gente):
        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.append(["123", "Mal", "Cuil", "", "", "", date(2026, 3, 1)])

        resultado = planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert resultado.creados == 0
        assert any("CUIL" in observacion for observacion in resultado.observaciones)

    def test_materia_desconocida_se_ignora_y_se_avisa(self, escuela_con_gente):
        libro = planilla.exportar(escuela_con_gente["institucion"])
        hoja = libro.active
        hoja.cell(row=2, column=14, value="Química | Alquimia")

        resultado = planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        persona = Legajo.objects.get(pk=escuela_con_gente["persona"].pk)
        assert list(persona.materias_que_puede_dar.all()) == [escuela_con_gente["quimica"]]
        assert any("Alquimia" in observacion for observacion in resultado.observaciones)

    def test_nadie_se_borra_por_faltar_en_el_archivo(self, escuela_con_gente):
        from openpyxl import Workbook

        libro = Workbook()
        libro.active.append(planilla.ENCABEZADOS)  # planilla vacía

        planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert Legajo.objects.filter(pk=escuela_con_gente["persona"].pk).exists()

    def test_marcar_de_baja_desde_la_planilla(self, escuela_con_gente):
        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.cell(row=2, column=12, value="De baja")

        planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        persona = Legajo.objects.get(pk=escuela_con_gente["persona"].pk)
        assert persona.estado == EstadoLegajo.BAJA


class TestElPlantelEnLaPlanilla:
    def test_va_y_vuelve_sin_cambios(self, escuela_con_gente):
        from legajos.models import Plantel

        persona = escuela_con_gente["persona"]
        persona.plantel = Plantel.PRECEPTOR
        persona.save()

        archivo = como_archivo(planilla.exportar(escuela_con_gente["institucion"]))
        resultado = planilla.importar(escuela_con_gente["institucion"], archivo)

        assert not resultado.observaciones
        assert Legajo.objects.get(pk=persona.pk).plantel == Plantel.PRECEPTOR

    def test_se_entiende_escrito_como_sea(self, escuela_con_gente):
        """«Ordenanza» a secas también vale: es como lo dice la escuela."""
        from legajos.models import Plantel

        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.cell(row=2, column=13, value="Ordenanza")

        planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        persona = Legajo.objects.get(pk=escuela_con_gente["persona"].pk)
        assert persona.plantel == Plantel.MAESTRANZA

    def test_un_plantel_desconocido_se_observa_y_no_toca(self, escuela_con_gente):
        from legajos.models import Plantel

        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.cell(row=2, column=13, value="Astronauta")

        resultado = planilla.importar(escuela_con_gente["institucion"], como_archivo(libro))

        assert any("Astronauta" in observacion for observacion in resultado.observaciones)
        assert Legajo.objects.get(pk=escuela_con_gente["persona"].pk).plantel == Plantel.DOCENTE


class TestLasPantallas:
    def test_descarga_un_xlsx(self, client, escuela_con_gente, secretaria):
        client.force_login(secretaria)
        respuesta = client.get(reverse("exportar_personal"))
        assert respuesta.status_code == 200
        assert "spreadsheetml" in respuesta["Content-Type"]

    def test_subir_la_planilla_desde_la_pantalla(self, client, escuela_con_gente, secretaria):
        client.force_login(secretaria)
        libro = planilla.exportar(escuela_con_gente["institucion"])
        libro.active.append(
            ["20-55666777-8", "Subida", "PorPantalla", "", "", "", date(2026, 3, 1)]
        )
        archivo = como_archivo(libro)
        archivo.name = "personal.xlsx"

        respuesta = client.post(reverse("importar_personal"), {"archivo": archivo})

        assert respuesta.status_code == 200
        assert Legajo.objects.filter(cuil="20-55666777-8").exists()
