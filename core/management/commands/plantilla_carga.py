"""La planilla para que una escuela nueva entregue sus datos.

Cargar una escuela es juntar ocho o nueve cosas —la estructura, la gente, los
cargos, el horario— que hoy viven en carpetas, en un Excel viejo y en la
cabeza de la secretaria. Este comando arma **un solo archivo** con una hoja
por cosa, las columnas que el sistema espera, desplegables con los valores
válidos y **una fila de ejemplo ya completada**, para que se complete sin
adivinar.

    python manage.py plantilla_carga
    python manage.py plantilla_carga --destino /tmp/escuela.xlsx

Las filas de ejemplo empiezan con la palabra ``EJEMPLO`` (ver
``core.planillas``): quien complete puede borrarlas o dejarlas, porque los
importadores las saltean solas.

Se completa en el orden de las hojas: cada una se apoya en la anterior. Ver
``MIGRAR-UNA-ESCUELA.md`` para el circuito completo.
"""

from dataclasses import dataclass, field
from pathlib import Path

from django.core.management.base import BaseCommand

from estructura.models import TipoBloque, TipoNivel, Vigencia
from horarios.models import MotivoNoDisponible
from legajos.models import FuentePago, MotivoBaja, Plantel, SituacionRevista, TipoCargo
from legajos.planilla import EJEMPLOS as EJEMPLOS_PERSONAL
from legajos.planilla import ENCABEZADOS as ENCABEZADOS_PERSONAL
from licencias.models import EstadoLicencia

VERDE = "0F6B63"
AMARILLO = "FFF8E1"
OCRE = "8A6D3B"

# El ejemplo va marcado para que ningún importador lo tome por un dato real.
EJ = "EJEMPLO "


def _etiquetas(choices) -> list[str]:
    """Los valores como los ve la persona que completa, no los internos."""
    return [etiqueta for _valor, etiqueta in choices]


DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
SI_NO = ["Sí", "No"]


@dataclass
class Hoja:
    titulo: str
    ayuda: str
    columnas: list[str]
    anchos: list[int]
    ejemplos: list[list]
    listas: dict[str, list[str]] = field(default_factory=dict)


HOJAS = [
    Hoja(
        titulo="Escuela",
        ayuda="Los datos de la institución. Una sola fila.",
        columnas=[
            "Nombre",
            "Nombre corto",
            "CUE",
            "CUIT",
            "Jurisdicción",
            "Domicilio",
            "Localidad",
            "Teléfono",
            "Email",
            "Color (ej. #1B5E20)",
        ],
        anchos=[38, 22, 14, 16, 16, 26, 18, 16, 28, 18],
        ejemplos=[
            [
                EJ + "Instituto Nuestra Señora del Carmen",
                "Nuestra Sra. del Carmen",
                "740012300",
                "30-71234567-8",
                "San Luis",
                "Av. Belgrano 1234",
                "Villa Mercedes",
                "2657 44-1234",
                "secretaria@carmen.edu.ar",
                "#1B5E20",
            ]
        ],
    ),
    Hoja(
        titulo="Niveles",
        ayuda="Los niveles que tiene la escuela. Una fila por nivel.",
        columnas=["Nivel", "Observaciones"],
        anchos=[18, 46],
        listas={"Nivel": _etiquetas(TipoNivel.choices)},
        ejemplos=[
            [EJ + "Primario", "Jornada simple, turno mañana y turno tarde"],
            [EJ + "Secundario", "Orientación en Economía y Administración"],
        ],
    ),
    Hoja(
        titulo="Ciclo y períodos",
        ayuda=(
            "El año escolar y sus cuatrimestres o trimestres. El horario se arma "
            "por período, así que el año se repite en cada fila."
        ),
        columnas=[
            "Año",
            "Inicio de clases",
            "Fin de clases",
            "Período",
            "Orden",
            "Desde",
            "Hasta",
        ],
        anchos=[10, 16, 16, 24, 8, 14, 14],
        ejemplos=[
            [
                EJ + "2026",
                "02/03/2026",
                "18/12/2026",
                "Primer cuatrimestre",
                1,
                "02/03/2026",
                "17/07/2026",
            ],
            [
                EJ + "2026",
                "02/03/2026",
                "18/12/2026",
                "Segundo cuatrimestre",
                2,
                "03/08/2026",
                "18/12/2026",
            ],
        ],
    ),
    Hoja(
        titulo="Turnos",
        ayuda="Los turnos de cada nivel, con su horario de entrada y de salida.",
        columnas=["Nivel", "Turno", "Entrada", "Salida"],
        anchos=[16, 18, 12, 12],
        listas={"Nivel": _etiquetas(TipoNivel.choices)},
        ejemplos=[
            [EJ + "Secundario", "Mañana", "07:30", "12:50"],
            [EJ + "Secundario", "Tarde", "13:30", "18:00"],
        ],
    ),
    Hoja(
        titulo="Grilla horaria",
        ayuda=(
            "Hora por hora, incluidos recreos y almuerzo. Un «esquema» es una grilla: "
            "si hay cursos que almuerzan y otros que no, son dos esquemas. Se repite "
            "para cada día, porque no todos los días son iguales."
        ),
        columnas=["Turno", "Esquema", "Día", "Orden", "Tipo", "Etiqueta", "Desde", "Hasta"],
        anchos=[16, 20, 14, 8, 18, 18, 12, 12],
        listas={"Día": DIAS, "Tipo": _etiquetas(TipoBloque.choices)},
        ejemplos=[
            [
                EJ + "Mañana",
                "Con almuerzo",
                "Lunes",
                1,
                "Hora de clase",
                "1ª hora",
                "07:30",
                "08:10",
            ],
            [
                EJ + "Mañana",
                "Con almuerzo",
                "Lunes",
                2,
                "Hora de clase",
                "2ª hora",
                "08:10",
                "08:50",
            ],
            [EJ + "Mañana", "Con almuerzo", "Lunes", 3, "Recreo", "Recreo", "08:50", "09:05"],
        ],
    ),
    Hoja(
        titulo="Cursos",
        ayuda="Los cursos y divisiones del ciclo. Ej.: 1° A, turno Mañana.",
        columnas=["Nivel", "Año de estudio", "División", "Turno", "Esquema"],
        anchos=[16, 15, 12, 16, 20],
        listas={"Nivel": _etiquetas(TipoNivel.choices)},
        ejemplos=[
            [EJ + "Secundario", 1, "A", "Mañana", "Con almuerzo"],
            [EJ + "Secundario", 1, "B", "Mañana", "Con almuerzo"],
        ],
    ),
    Hoja(
        titulo="Materias",
        ayuda=(
            "Las materias de cada nivel. El nombre se escribe una sola vez acá y "
            "después se repite igual en el plan de estudios y en los cargos."
        ),
        columnas=["Nivel", "Materia", "Abreviatura"],
        anchos=[16, 34, 14],
        listas={"Nivel": _etiquetas(TipoNivel.choices)},
        ejemplos=[
            [EJ + "Secundario", "Matemática", "Mat"],
            [EJ + "Secundario", "Lengua y Literatura", "Len"],
        ],
    ),
    Hoja(
        titulo="Plan de estudios",
        ayuda=(
            "Qué materias tiene cada curso y con cuántas horas semanales. La columna "
            "«Período» solo se completa si la materia dura un período nada más."
        ),
        columnas=["Curso (ej. 1°A)", "Materia", "Horas semanales", "Vigencia", "Período"],
        anchos=[18, 34, 16, 20, 22],
        listas={"Vigencia": _etiquetas(Vigencia.choices)},
        ejemplos=[
            [EJ + "1°A", "Matemática", 5, "Todo el año", ""],
            [EJ + "1°A", "Educación Física", 2, "Solo un período", "Primer cuatrimestre"],
        ],
    ),
    Hoja(
        titulo="Personal",
        ayuda=(
            "Una fila por persona, docente y no docente. El CUIL identifica: con el "
            "mismo se actualiza, con uno nuevo se crea. Alcanza con CUIL, apellido, "
            "nombre y fecha de ingreso; el resto se puede completar después."
        ),
        columnas=ENCABEZADOS_PERSONAL,
        anchos=[22, 22, 22, 12, 30, 16, 15, 16, 18, 26, 16, 12, 22, 36],
        listas={
            "Estado": ["Activo", "De baja"],
            "Plantel": _etiquetas(Plantel.choices),
        },
        ejemplos=EJEMPLOS_PERSONAL,
    ),
    Hoja(
        titulo="Cargos",
        ayuda=(
            "Una fila por cargo: alguien con 15 horas en tres cursos son tres filas. "
            "La fuente de pago es lo que define a qué planilla va cada novedad, y por "
            "eso una misma persona puede tener cargos de las dos. El apellido y el "
            "nombre van para poder identificar a la persona mientras no haya CUIL."
        ),
        columnas=[
            "CUIL",
            "Apellido",
            "Nombre",
            "Tipo de cargo",
            "Denominación",
            "Nivel",
            "Materia",
            "Curso (ej. 1°A)",
            "Horas semanales",
            "Jornada completa",
            "Situación de revista",
            "Fuente de pago",
            "Fecha de alta",
            "Fecha de baja",
            "Motivo de baja",
            "Resolución n°",
            "Fecha de resolución",
        ],
        anchos=[22, 22, 22, 24, 26, 14, 26, 16, 15, 16, 22, 32, 14, 14, 18, 16, 18],
        listas={
            "Tipo de cargo": _etiquetas(TipoCargo.choices),
            "Nivel": _etiquetas(TipoNivel.choices),
            "Jornada completa": SI_NO,
            "Situación de revista": _etiquetas(SituacionRevista.choices),
            "Fuente de pago": _etiquetas(FuentePago.choices),
            "Motivo de baja": _etiquetas(MotivoBaja.choices),
        },
        ejemplos=[
            [
                EJ + "27-30111222-4",
                "Benítez",
                "María Laura",
                "Horas cátedra (40 min)",
                "Profesora de Matemática",
                "Secundario",
                "Matemática",
                "1°A",
                5,
                "No",
                "Titular",
                "Subvencionado (lo paga el estado)",
                "01/03/2018",
                "",
                "",
                "1234-ME-2018",
                "20/02/2018",
            ],
            [
                EJ + "27-30111222-4",
                "Benítez",
                "María Laura",
                "Horas cátedra (40 min)",
                "Profesora de Matemática",
                "Secundario",
                "Matemática",
                "2°A",
                4,
                "No",
                "Suplente",
                "Interno (lo paga la escuela)",
                "02/03/2026",
                "",
                "",
                "",
                "",
            ],
        ],
    ),
    Hoja(
        titulo="Licencias",
        ayuda=(
            "Las licencias del año en curso, si se quieren tener cargadas. Opcional: "
            "el sistema funciona igual sin el historial."
        ),
        columnas=["CUIL", "Artículo (ej. Art. 76)", "Desde", "Hasta", "Estado", "Observaciones"],
        anchos=[22, 22, 14, 14, 16, 46],
        listas={"Estado": _etiquetas(EstadoLicencia.choices)},
        ejemplos=[
            [
                EJ + "27-30111222-4",
                "Art. 76",
                "05/05/2026",
                "07/05/2026",
                "Aprobada",
                "Enfermedad de corta duración. Certificado presentado.",
            ]
        ],
    ),
    Hoja(
        titulo="Documentación",
        ayuda=(
            "Vencimientos que la escuela controla (apto psicofísico, antecedentes "
            "penales). Opcional. Los archivos escaneados se suben después, uno por uno."
        ),
        columnas=["CUIL", "Documento", "Fecha de emisión", "Fecha de vencimiento"],
        anchos=[22, 34, 18, 20],
        ejemplos=[[EJ + "27-30111222-4", "Apto psicofísico", "15/02/2026", "15/02/2027"]],
    ),
    Hoja(
        titulo="Disponibilidad (DDJJ)",
        ayuda=(
            "En qué franjas cada docente NO puede tomar horas. Solo hace falta si se "
            "quiere que el sistema genere el horario. Opcional."
        ),
        columnas=["CUIL", "Período", "Día", "Desde", "Hasta", "Motivo", "¿Es preferencia?"],
        anchos=[22, 22, 14, 12, 12, 32, 18],
        listas={
            "Día": DIAS,
            "Motivo": _etiquetas(MotivoNoDisponible.choices),
            "¿Es preferencia?": SI_NO,
        },
        ejemplos=[
            [
                EJ + "27-30111222-4",
                "Primer cuatrimestre",
                "Martes",
                "07:30",
                "12:50",
                "Trabaja en otra institución",
                "No",
            ]
        ],
    ),
]

INSTRUCCIONES = [
    ("Para qué es esta planilla", ""),
    (
        "",
        "Es todo lo que el sistema necesita para empezar a funcionar con los datos "
        "de una escuela de verdad. Una hoja por cosa, en el orden en que hay que "
        "completarlas: cada una se apoya en la anterior.",
    ),
    ("", ""),
    ("Cómo completarla", ""),
    ("", "1. Completá las hojas en orden, de izquierda a derecha."),
    (
        "",
        "2. Cada hoja trae una o dos filas amarillas de ejemplo, ya completadas, para "
        "que se vea la forma de cada dato. Escribí abajo de ellas. Podés borrarlas o "
        "dejarlas: empiezan con la palabra EJEMPLO y el sistema las saltea solo.",
    ),
    (
        "",
        "3. Las celdas con desplegable solo aceptan los valores de la lista. "
        "Si algo no encaja en ninguno, dejalo vacío y anotalo aparte: lo vemos.",
    ),
    ("", "4. Las fechas van como fecha de Excel o escritas dd/mm/aaaa."),
    (
        "",
        "5. Los nombres se escriben una sola vez y después se repiten IGUAL. "
        "«Matemática» y «Matematica» son dos materias distintas para la computadora.",
    ),
    (
        "",
        "6. Los cursos se escriben siempre igual en todas las hojas: 1°A, 1°B, 2°A…",
    ),
    ("", ""),
    ("Qué es obligatorio y qué no", ""),
    (
        "",
        "Imprescindible: Escuela, Niveles, Ciclo y períodos, Turnos, Grilla horaria, "
        "Cursos, Materias, Plan de estudios, Personal y Cargos.",
    ),
    (
        "",
        "Opcional: Licencias, Documentación y Disponibilidad. Sirven para arrancar "
        "con el historial cargado, pero el sistema funciona sin eso.",
    ),
    ("", ""),
    ("Sobre los datos personales", ""),
    (
        "",
        "Acá hay CUIL, teléfonos y domicilios de personas reales. Para probar el "
        "sistema alcanza con apellido, nombre, CUIL y los cargos: el resto puede "
        "esperar. Compartí el archivo solo con quien tenga que verlo.",
    ),
    ("", ""),
    ("Si algo no se entiende", ""),
    (
        "",
        "Es mejor dejar la celda vacía y preguntar que inventar un dato. Un dato "
        "inventado en una planilla de sueldos es plata de más o de menos en el "
        "sueldo de alguien.",
    ),
]

FILA_ENCABEZADO = 3


class Command(BaseCommand):
    help = "Genera el Excel para que una escuela nueva entregue sus datos."

    def add_arguments(self, parser):
        parser.add_argument(
            "--destino",
            default="plantilla-carga-escuela.xlsx",
            help="Dónde guardar el archivo.",
        )

    def handle(self, *args, **opciones):
        from openpyxl import Workbook

        libro = Workbook()
        self._instrucciones(libro.active)
        for hoja in HOJAS:
            self._armar(libro.create_sheet(hoja.titulo[:31]), hoja)

        destino = Path(opciones["destino"])
        libro.save(destino)

        self.stdout.write(self.style.SUCCESS(f"Planilla lista: {destino.resolve()}"))
        self.stdout.write(f"  {len(HOJAS)} hojas: {', '.join(h.titulo for h in HOJAS)}")
        self.stdout.write(
            "  Cada hoja trae su fila de ejemplo; se puede dejar, porque empieza con "
            "«EJEMPLO» y los importadores la saltean."
        )
        self.stdout.write(
            "  Las hojas de Licencias, Documentación y Disponibilidad son opcionales."
        )

    def _armar(self, hoja, definicion: Hoja):
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # La ayuda va arriba de todo, donde se lee antes de empezar.
        hoja["A1"] = definicion.ayuda
        hoja["A1"].font = Font(italic=True, color="4A5C6B", size=10)
        hoja["A1"].alignment = Alignment(vertical="center")
        hoja.append([])
        hoja.append(definicion.columnas)
        for celda in hoja[FILA_ENCABEZADO]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor=VERDE)

        for ejemplo in definicion.ejemplos:
            hoja.append(ejemplo)
            for celda in hoja[hoja.max_row]:
                celda.font = Font(italic=True, color=OCRE)
                celda.fill = PatternFill("solid", fgColor=AMARILLO)

        for indice, ancho in enumerate(definicion.anchos, start=1):
            hoja.column_dimensions[get_column_letter(indice)].width = ancho
        hoja.freeze_panes = f"A{FILA_ENCABEZADO + 1}"

        # Los desplegables arrancan debajo del ejemplo: si abarcaran la fila de
        # muestra, Excel la marcaría como valor inválido y asustaría al que completa.
        primera = FILA_ENCABEZADO + len(definicion.ejemplos) + 1
        for columna, valores in definicion.listas.items():
            letra = get_column_letter(definicion.columnas.index(columna) + 1)
            regla = DataValidation(
                type="list",
                formula1='"' + ",".join(valores) + '"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valor no válido",
                error="Elegí uno de los valores de la lista.",
            )
            hoja.add_data_validation(regla)
            regla.add(f"{letra}{primera}:{letra}500")

    def _instrucciones(self, hoja):
        from openpyxl.styles import Alignment, Font

        hoja.title = "Instrucciones"
        hoja.column_dimensions["A"].width = 4
        hoja.column_dimensions["B"].width = 100

        hoja["B1"] = "Carga de datos de la escuela"
        hoja["B1"].font = Font(bold=True, size=14)

        fila = 3
        for titulo, texto in INSTRUCCIONES:
            if titulo:
                hoja[f"B{fila}"] = titulo
                hoja[f"B{fila}"].font = Font(bold=True, size=11)
            elif texto:
                hoja[f"B{fila}"] = texto
                hoja[f"B{fila}"].alignment = Alignment(wrap_text=True, vertical="top")
                hoja.row_dimensions[fila].height = 30
            fila += 1
