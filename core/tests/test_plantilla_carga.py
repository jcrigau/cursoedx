"""La planilla que se le manda a una escuela nueva.

Lo que importa: que las columnas y los ejemplos coincidan, que los ejemplos
usen valores que el desplegable acepta —si el ejemplo miente, la escuela
copia la mentira— y que queden marcados para que ningún importador los tome
por datos reales.
"""

import pytest
from django.core.management import call_command
from openpyxl import load_workbook

from core.management.commands.plantilla_carga import FILA_ENCABEZADO, HOJAS
from core.planillas import es_ejemplo


class TestLasHojas:
    def test_cada_ejemplo_tiene_tantas_celdas_como_columnas(self):
        for hoja in HOJAS:
            for ejemplo in hoja.ejemplos:
                assert len(ejemplo) == len(hoja.columnas), hoja.titulo

    def test_hay_un_ancho_por_columna(self):
        for hoja in HOJAS:
            assert len(hoja.anchos) == len(hoja.columnas), hoja.titulo

    def test_todas_las_hojas_muestran_un_ejemplo(self):
        for hoja in HOJAS:
            assert hoja.ejemplos, hoja.titulo

    def test_los_ejemplos_van_marcados(self):
        for hoja in HOJAS:
            for ejemplo in hoja.ejemplos:
                assert es_ejemplo(ejemplo[0]), f"{hoja.titulo}: {ejemplo[0]}"

    def test_el_ejemplo_usa_los_valores_del_desplegable(self):
        for hoja in HOJAS:
            for columna, valores in hoja.listas.items():
                indice = hoja.columnas.index(columna)
                if indice == 0:
                    continue  # esa celda lleva la marca de ejemplo adelante
                for ejemplo in hoja.ejemplos:
                    valor = ejemplo[indice]
                    assert valor in valores or valor == "", f"{hoja.titulo} / {columna}: «{valor}»"

    def test_las_columnas_con_desplegable_existen(self):
        for hoja in HOJAS:
            for columna in hoja.listas:
                assert columna in hoja.columnas, f"{hoja.titulo}: {columna}"


class TestElArchivo:
    @pytest.fixture
    def libro(self, tmp_path):
        destino = tmp_path / "plantilla.xlsx"
        call_command("plantilla_carga", destino=str(destino))
        return load_workbook(destino)

    def test_trae_las_instrucciones_y_una_hoja_por_cosa(self, libro):
        assert libro.sheetnames == ["Instrucciones"] + [h.titulo for h in HOJAS]

    def test_los_encabezados_estan_donde_se_esperan(self, libro):
        for definicion in HOJAS:
            hoja = libro[definicion.titulo]
            fila = [celda.value for celda in hoja[FILA_ENCABEZADO]]
            assert fila[: len(definicion.columnas)] == definicion.columnas

    def test_el_ejemplo_arranca_debajo_del_encabezado(self, libro):
        for definicion in HOJAS:
            hoja = libro[definicion.titulo]
            primera = hoja.cell(row=FILA_ENCABEZADO + 1, column=1).value
            assert es_ejemplo(primera), definicion.titulo

    def test_el_desplegable_no_abarca_la_fila_de_ejemplo(self, libro):
        for definicion in HOJAS:
            if not definicion.listas:
                continue
            hoja = libro[definicion.titulo]
            primera_libre = FILA_ENCABEZADO + len(definicion.ejemplos) + 1
            for regla in hoja.data_validations.dataValidation:
                assert str(regla.sqref).split(":")[0][1:] == str(primera_libre)
