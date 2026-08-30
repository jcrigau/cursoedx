"""La planilla de carga entra entera: estructura, personal y cargos.

Lo que importa: que cada hoja se apoye en la anterior y lo diga cuando falta,
que nada se cargue a medias —un cargo mal cargado es plata— y que volver a
correrlo actualice en vez de duplicar. Una escuela se migra corrigiendo el
Excel y reintentando cinco veces; si la segunda corrida duplicara, no se
podría.
"""

import pytest
from django.core.management import call_command
from openpyxl import load_workbook

from core.models import Institucion
from core.planillas import FILA_ENCABEZADO
from estructura.models import Curso, Materia, Nivel, Turno
from legajos.models import Cargo, FuentePago, Legajo

HORARIO = [("07:30", "08:10"), ("08:10", "08:50"), ("09:05", "09:45")]


@pytest.fixture
def planilla(tmp_path):
    """La plantilla en blanco, con un ayudante para completar hojas."""
    destino = tmp_path / "escuela.xlsx"
    call_command("plantilla_carga", destino=str(destino))
    libro = load_workbook(destino)

    def completar(hoja, filas):
        h = libro[hoja]
        if h.max_row > FILA_ENCABEZADO:
            h.delete_rows(FILA_ENCABEZADO + 1, h.max_row - FILA_ENCABEZADO)
        for fila in filas:
            h.append(fila)

    def guardar():
        libro.save(destino)
        return str(destino)

    completar.guardar = guardar
    completar.libro = libro
    return completar


def escuela_minima(planilla, *, con_cargos=True):
    planilla("Escuela", [["Escuela de Prueba", "Prueba", "", "", "San Luis"]])
    planilla("Niveles", [["Secundario", ""]])
    planilla(
        "Ciclo y períodos",
        [[2026, "02/03/2026", "18/12/2026", "Primer cuatrimestre", 1, "02/03/2026", "17/07/2026"]],
    )
    planilla("Turnos", [["Secundario", "Mañana", "07:30", "12:50"]])
    planilla(
        "Grilla horaria",
        [
            ["Mañana", "Único", "Lunes", orden, "Hora de clase", f"{orden}ª", desde, hasta]
            for orden, (desde, hasta) in enumerate(HORARIO, start=1)
        ],
    )
    planilla("Cursos", [["Secundario", 1, "A", "", ""], ["Secundario", 1, "B", "", ""]])
    planilla("Materias", [["Secundario", "Matemática", "Mat"], ["Secundario", "Lengua", "Len"]])
    planilla("Plan de estudios", [["1°A", "Matemática", 5, "Todo el año", ""]])
    planilla(
        "Personal",
        [
            [
                "27-30111222-4",
                "Benítez",
                "Ana",
                "",
                "",
                "",
                "01/03/2020",
                "",
                "",
                "",
                "",
                "Activo",
                "Docente",
                "Matemática",
            ]
        ],
    )
    planilla(
        "Cargos",
        [
            [
                "27-30111222-4",
                "Benítez",
                "Ana",
                "Horas cátedra (40 min)",
                "",
                "Secundario",
                "Matemática",
                "1°A",
                5,
                "No",
                "Titular",
                "Subvencionado (lo paga el estado)",
                "01/03/2020",
                "",
                "",
                "R-1",
                "",
            ]
        ]
        if con_cargos
        else [],
    )
    return planilla.guardar()


class TestLaCargaCompleta:
    def test_entra_la_escuela_entera(self, db, planilla, capsys):
        call_command("cargar_planilla", escuela_minima(planilla))

        institucion = Institucion.objects.get(nombre="Escuela de Prueba")
        assert Nivel.objects.filter(institucion=institucion).count() == 1
        assert Turno.objects.filter(institucion=institucion).count() == 1
        assert Curso.objects.filter(institucion=institucion).count() == 2
        assert Materia.objects.filter(institucion=institucion).count() == 2
        assert Legajo.objects.filter(institucion=institucion).count() == 1
        cargo = Cargo.objects.get(institucion=institucion)
        assert cargo.horas_semanales == 5
        assert cargo.fuente_pago == FuentePago.SUBVENCIONADO
        assert str(cargo.curso) == "1°A"

    def test_volver_a_correrlo_no_duplica(self, db, planilla):
        archivo = escuela_minima(planilla)
        call_command("cargar_planilla", archivo)
        call_command("cargar_planilla", archivo)

        assert Curso.objects.count() == 2
        assert Legajo.objects.count() == 1
        assert Cargo.objects.count() == 1

    def test_simular_no_guarda_nada(self, db, planilla):
        call_command("cargar_planilla", escuela_minima(planilla), simular=True)

        assert not Institucion.objects.exists()
        assert not Legajo.objects.exists()

    def test_el_curso_hereda_el_turno_cuando_hay_uno_solo(self, db, planilla):
        call_command("cargar_planilla", escuela_minima(planilla))

        assert Curso.objects.get(division="A").turno.nombre == "Mañana"

    def test_queda_en_la_bitacora(self, db, planilla):
        from core.models import RegistroAuditoria

        call_command("cargar_planilla", escuela_minima(planilla))

        assert RegistroAuditoria.objects.filter(accion="IMPORTACION").exists()


class TestCuandoFaltaAlgo:
    def test_sin_ciclo_no_hay_cursos_y_lo_explica(self, db, planilla, capsys):
        escuela_minima(planilla)
        planilla("Ciclo y períodos", [])
        call_command("cargar_planilla", planilla.guardar())

        salida = capsys.readouterr().out
        assert not Curso.objects.exists()
        assert "no hay ciclo lectivo" in salida

    def test_el_cargo_de_una_materia_que_no_existe_no_entra(self, db, planilla, capsys):
        escuela_minima(planilla)
        planilla(
            "Cargos",
            [
                [
                    "27-30111222-4",
                    "Benítez",
                    "Ana",
                    "Horas cátedra (40 min)",
                    "",
                    "Secundario",
                    "Alquimia",
                    "1°A",
                    5,
                    "No",
                    "Titular",
                    "Subvencionado (lo paga el estado)",
                    "01/03/2020",
                    "",
                    "",
                    "",
                    "",
                ]
            ],
        )
        call_command("cargar_planilla", planilla.guardar())

        assert not Cargo.objects.exists()
        assert "Alquimia" in capsys.readouterr().out

    def test_un_cargo_sin_fuente_de_pago_no_entra(self, db, planilla, capsys):
        escuela_minima(planilla)
        planilla(
            "Cargos",
            [
                [
                    "27-30111222-4",
                    "Benítez",
                    "Ana",
                    "Horas cátedra (40 min)",
                    "",
                    "Secundario",
                    "Matemática",
                    "1°A",
                    5,
                    "No",
                    "Titular",
                    "",
                    "01/03/2020",
                    "",
                    "",
                    "",
                    "",
                ]
            ],
        )
        call_command("cargar_planilla", planilla.guardar())

        assert not Cargo.objects.exists()
        assert "fuente de pago" in capsys.readouterr().out

    def test_el_curso_combinado_se_explica(self, db, planilla, capsys):
        escuela_minima(planilla)
        planilla(
            "Cargos",
            [
                [
                    "27-30111222-4",
                    "Benítez",
                    "Ana",
                    "Horas cátedra (40 min)",
                    "",
                    "Secundario",
                    "Matemática",
                    "1°AB",
                    5,
                    "No",
                    "Titular",
                    "Subvencionado (lo paga el estado)",
                    "01/03/2020",
                    "",
                    "",
                    "",
                    "",
                ]
            ],
        )
        call_command("cargar_planilla", planilla.guardar())

        assert not Cargo.objects.exists()
        assert "curso combinado" in capsys.readouterr().out

    def test_la_plantilla_en_blanco_no_crea_una_escuela_de_ejemplo(self, db, planilla):
        """Las filas de muestra son muestra: no pueden fundar el «Instituto Ejemplo»."""
        from django.core.management.base import CommandError

        with pytest.raises(CommandError, match="Escuela"):
            call_command("cargar_planilla", planilla.guardar())

        assert not Institucion.objects.exists()


class TestLosCargosSinCuil:
    def test_se_cuelgan_de_la_persona_por_apellido_y_nombre(self, db, planilla):
        escuela_minima(planilla)
        planilla(
            "Personal",
            [
                [
                    "",
                    "Ochoa",
                    "Ramiro",
                    "",
                    "",
                    "",
                    "01/03/2020",
                    "",
                    "",
                    "",
                    "",
                    "Activo",
                    "Docente",
                    "Matemática",
                ]
            ],
        )
        planilla(
            "Cargos",
            [
                [
                    "",
                    "Ochoa",
                    "Ramiro",
                    "Horas cátedra (40 min)",
                    "",
                    "Secundario",
                    "Matemática",
                    "1°A",
                    4,
                    "No",
                    "Titular",
                    "Interno (lo paga la escuela)",
                    "01/03/2020",
                    "",
                    "",
                    "",
                    "",
                ]
            ],
        )
        call_command("cargar_planilla", planilla.guardar())

        cargo = Cargo.objects.get()
        assert cargo.legajo.apellido == "Ochoa"
        assert cargo.legajo.cuil == ""
        assert cargo.fuente_pago == FuentePago.INTERNO
